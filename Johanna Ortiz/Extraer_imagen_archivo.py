import os
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image

# Utilidades
HEADER_CODIGO = "CODIGO OT"
HEADER_IMAGEN = "IMAGEN"

def normaliza_header(s):
    return str(s).strip().upper().replace("\n", " ") if s is not None else ""

def es_codigo_valido(codigo):
    if codigo is None:
        return False
    s = str(codigo).strip()
    if not s:
        return False
    return s.upper() not in {"N/A", "NA", "NONE"}
    
def limpia_nombre_archivo(s):
    s = str(s).strip()
    return "".join(c for c in s if c.isalnum() or c in ("-", "_"))

def detectar_fila_y_indices_headers(ws, max_filas_busqueda=10):
    """
    Busca en las primeras 'max_filas_busqueda' filas una fila que contenga los headers
    requeridos. Devuelve (fila_header, idx_codigo, idx_imagen) en base 1.
    Si no encuentra, retorna (None, None, None).
    """
    for row in ws.iter_rows(min_row=1, max_row=max_filas_busqueda, values_only=True):
        if row is None:
            continue
        encabezados_norm = [normaliza_header(c) for c in row]
        if HEADER_CODIGO in encabezados_norm and HEADER_IMAGEN in encabezados_norm:
            idx_codigo = encabezados_norm.index(HEADER_CODIGO) + 1
            idx_imagen = encabezados_norm.index(HEADER_IMAGEN) + 1
            # fila_header = índice real de esa fila (iter_rows no da índice directamente),
            # así que lo calculamos por longitud de lista acumulada. Más simple: usamos
            # ws.iter_rows con enumerate para conocer la fila.
            # Recorremos de nuevo de forma controlada:
            # (optamos por la forma explícita para precisión)
            fila_header = None
            r = 1
            for rr in ws.iter_rows(min_row=1, max_row=max_filas_busqueda, values_only=True):
                if rr == row:
                    fila_header = r
                    break
                r += 1
            return fila_header, idx_codigo, idx_imagen
    return None, None, None

def exportar_imagenes_excel(ruta_excel, carpeta_salida="imagenes"):
    os.makedirs(carpeta_salida, exist_ok=True)

    wb = load_workbook(ruta_excel, data_only=True)

    # Cache de índices de columnas por hoja: {nombre_hoja: (fila_header, idx_codigo, idx_imagen)}
    headers_hojas = {}
    hojas_validas = []

    # Pre-indexar todas las hojas que tengan ambos encabezados
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        fila_header, idx_codigo, idx_imagen = detectar_fila_y_indices_headers(ws)
        if fila_header and idx_codigo and idx_imagen:
            headers_hojas[nombre_hoja] = (fila_header, idx_codigo, idx_imagen)
            hojas_validas.append(nombre_hoja)
        else:
            print(f"⚠️ En la hoja '{nombre_hoja}' no encontré columnas requeridas '{HEADER_CODIGO}' y '{HEADER_IMAGEN}' en las primeras filas.")

    if not hojas_validas:
        print("❌ No se encontraron hojas con los encabezados requeridos en el libro.")
        return

    # Recorremos todas las hojas (porque las imágenes pueden estar en cualquiera)
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]

        # Si esta hoja no tiene headers válidos, igual puede contener imágenes.
        # Intentaremos resolver el código consultando esta u otras hojas.
        # Si los tiene, genial: será el primer intento.
        tiene_headers_aqui = nombre_hoja in headers_hojas
        if tiene_headers_aqui:
            _, idx_codigo_local, _ = headers_hojas[nombre_hoja]
        else:
            idx_codigo_local = None

        # Recorremos imágenes insertadas en la hoja actual
        if not getattr(ws, "_images", None):
            # Sin imágenes; continuamos
            continue

        for image in ws._images:
            # Fila (base 1) donde está anclada la imagen
            try:
                celda = image.anchor._from  # openpyxl < 3.1
                fila_img = celda.row + 1
            except AttributeError:
                # Compatibilidad con versiones nuevas: image.anchor (OneCellAnchor/TwoCellAnchor)
                # Intento seguro: obtener fila desde ._from si existe, si no, saltamos
                print(f"⚠️ No pude leer la fila de una imagen en hoja '{nombre_hoja}'. La salto.")
                continue

            codigo = None

            # 1) Intento leer código en la MISMA hoja (si tiene headers)
            if tiene_headers_aqui and fila_img > 0:
                valor = ws.cell(row=fila_img, column=idx_codigo_local).value
                if es_codigo_valido(valor):
                    codigo = valor

            # 2) Si no lo conseguí aquí, pruebo en el MISMO número de fila de las OTRAS hojas válidas
            if not es_codigo_valido(codigo):
                for otra_hoja in hojas_validas:
                    if otra_hoja == nombre_hoja:
                        continue
                    ws2 = wb[otra_hoja]
                    _, idx_codigo2, _ = headers_hojas[otra_hoja]
                    valor2 = ws2.cell(row=fila_img, column=idx_codigo2).value
                    if es_codigo_valido(valor2):
                        codigo = valor2
                        print(f"ℹ️ Código para imagen en '{nombre_hoja}' fila {fila_img} obtenido desde hoja '{otra_hoja}'.")
                        break

            # 3) Si aún no hay código, no podemos nombrar el archivo de forma confiable
            if not es_codigo_valido(codigo):
                print(f"⚠️ Saltada imagen en hoja '{nombre_hoja}' fila {fila_img}: no encontré '{HEADER_CODIGO}' válido en esta ni en otras hojas.")
                continue

            # Limpieza de nombre
            codigo_str = limpia_nombre_archivo(codigo)

            # Obtener bytes de la imagen
            try:
                data = image._data()
            except Exception:
                print(f"⚠️ No pude extraer datos binarios de una imagen en hoja '{nombre_hoja}', fila {fila_img}.")
                continue

            # Detectar extensión con PIL
            try:
                with Image.open(BytesIO(data)) as img:
                    ext = img.format.lower() if img.format else "png"
            except Exception:
                ext = "png"  # fallback

            nombre_archivo = f"{codigo_str}.{ext}"
            ruta_guardado = os.path.join(carpeta_salida, nombre_archivo)

            with open(ruta_guardado, "wb") as f:
                f.write(data)

            print(f"✅ Guardada: {ruta_guardado}")

    print("🚀 Proceso terminado.")

# Uso:
exportar_imagenes_excel("Copia5.xlsx", "OT 5")
